# -*- coding: utf-8 -*-
"""
Eserler.xlsx dosyasından src/data/eserler.ts üretir.

Kullanım (proje kökünden):
    python scripts/eserler-uret.py

Excel dosyası proje klasörünün bir üstünde beklenir:
    SardunyaWebPage/Eserler.xlsx
    SardunyaWebPage/ilknur-gurcan/   <- proje kökü

Sütunlar başlık adıyla bulunur, sıraları değişebilir. Beklenen başlıklar:
    Dosya Adı, Boyutlar (En x Boy cm), Yıl, Kategori, Fiyat (TL), Durum,
    Seçili Resim mi?, Hero Resmi mi?, Yayınlansın mı?, Onaylanan İsim,
    İngilizce İsim

Not: eserler.ts elle düzenlenmemelidir; Excel güncellenip bu betik
çalıştırılmalıdır.
"""
import re
import sys
import unicodedata
from pathlib import Path

# Windows konsolu varsayılan olarak cp1254 kullanır; Türkçe olmayan
# karakterlerde (→ gibi) çökmemesi için çıktıyı UTF-8'e sabitle.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl kurulu değil.  Kurulum:  pip install openpyxl")

KOK  = Path(__file__).resolve().parent.parent
XLSX = KOK.parent / "Eserler.xlsx"
CIKTI = KOK / "src" / "data" / "eserler.ts"

# Türkçe harfleri slug için sadeleştir
TR_HARITA = str.maketrans({
    'ç': 'c', 'Ç': 'c', 'ğ': 'g', 'Ğ': 'g', 'ı': 'i', 'I': 'i',
    'İ': 'i', 'ö': 'o', 'Ö': 'o', 'ş': 's', 'Ş': 's', 'ü': 'u',
    'Ü': 'u', 'â': 'a', 'Â': 'a', 'î': 'i', 'û': 'u',
})


def slugify(metin: str) -> str:
    s = metin.translate(TR_HARITA).lower()
    s = re.sub(r"['’]", "", s)            # kesme işareti: sahaf'ta -> sahafta
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def boyut(deger) -> tuple[int, int]:
    sayilar = re.findall(r"\d+", str(deger))
    if len(sayilar) < 2:
        raise ValueError(f"Boyut okunamadı: {deger!r}")
    return int(sayilar[0]), int(sayilar[1])


def ts_metin(s: str) -> str:
    """TS string literal: kesme işareti varsa çift tırnak kullan."""
    return f'"{s}"' if "'" in s else f"'{s}'"


def sutunlar(ws) -> dict[str, int]:
    """
    Başlık adı -> sütun numarası.

    DİKKAT: Excel'de aynı başlık birden fazla sütunda olabilir
    (ör. 'Durum' hem satış durumu hem not sütunu için kullanılmış).
    Bu durumda İLK sütun geçerlidir; sonrakiler yok sayılır ve uyarı basılır.
    """
    harita: dict[str, int] = {}
    mukerrer: dict[str, list[str]] = {}

    for c in range(1, ws.max_column + 1):
        ham = ws.cell(row=1, column=c).value
        if not ham:
            continue
        ad = str(ham).strip()
        harf = openpyxl.utils.get_column_letter(c)
        if ad in harita:
            mukerrer.setdefault(ad, [openpyxl.utils.get_column_letter(harita[ad])]).append(harf)
            continue          # ilk sütun kazanır
        harita[ad] = c

    for ad, harfler in mukerrer.items():
        print(f"  UYARI: '{ad}' başlığı birden fazla sütunda ({', '.join(harfler)}). "
              f"{harfler[0]} sütunu kullanıldı.")

    return harita


def oku(ws, sut, satir, baslik, zorunlu=True):
    if baslik not in sut:
        if zorunlu:
            raise KeyError(f"Excel'de '{baslik}' sütunu yok")
        return None
    return ws.cell(row=satir, column=sut[baslik]).value


def main() -> int:
    if not XLSX.exists():
        sys.exit(f"Excel bulunamadı: {XLSX}")

    ws = openpyxl.load_workbook(XLSX).active
    sut = sutunlar(ws)

    eserler, slugler, atlanan = [], set(), 0

    for r in range(2, ws.max_row + 1):
        dosya = oku(ws, sut, r, "Dosya Adı")
        if not dosya:
            continue

        if str(oku(ws, sut, r, "Yayınlansın mı?")).strip().lower() != "evet":
            atlanan += 1
            continue

        baslik = (oku(ws, sut, r, "Onaylanan İsim") or "").strip()
        if not baslik:
            sys.exit(f"Satır {r}: 'Onaylanan İsim' boş ({dosya})")

        baslik_en = (oku(ws, sut, r, "İngilizce İsim", zorunlu=False) or baslik).strip()

        slug = slugify(baslik)
        if slug in slugler:                       # aynı isim iki kez ise ayır
            i = 2
            while f"{slug}-{i}" in slugler:
                i += 1
            slug = f"{slug}-{i}"
        slugler.add(slug)

        en, boy = boyut(oku(ws, sut, r, "Boyutlar (En x Boy cm)"))
        satildi = str(oku(ws, sut, r, "Durum")).strip().lower().startswith("satıld")
        fiyat = oku(ws, sut, r, "Fiyat (TL)")

        eserler.append({
            "slug": slug,
            "baslik": baslik,
            "baslik_en": baslik_en,
            "kategori": str(oku(ws, sut, r, "Kategori")).strip(),
            "yil": int(oku(ws, sut, r, "Yıl")),
            "en": en,
            "boy": boy,
            "fiyat": f'"{fiyat}"' if isinstance(fiyat, str) else str(int(fiyat or 0)),
            "durum": "satildi" if satildi else "satilabilir",
            "foto": f"/images/{dosya}",
            "secili": str(oku(ws, sut, r, "Seçili Resim mi?")).strip().lower().startswith("seçili"),
            "hero": str(oku(ws, sut, r, "Hero Resmi mi?")).strip().lower().startswith("hero"),
        })

    if not eserler:
        sys.exit("Yayınlanacak eser bulunamadı.")

    eserler.sort(key=lambda e: -e["yil"])          # en yeni önce

    # Sütunları hizala (dosya elle okunabilir kalsın)
    g = lambda alan, ek=2: max(len(ts_metin(e[alan])) for e in eserler) + ek
    w_slug, w_bas, w_ben = g("slug"), g("baslik"), g("baslik_en")
    w_kat, w_foto = g("kategori"), g("foto") - 1
    w_fiy = max(len(e["fiyat"]) for e in eserler) + 2

    satirlar = []
    for e in eserler:
        ek = ""
        if e["secili"]:
            ek += ", secili: true"
        if e["hero"]:
            ek += ", hero: true"
        satirlar.append(
            f"  {{ slug: {ts_metin(e['slug']) + ',':{w_slug}} "
            f"baslik: {ts_metin(e['baslik']) + ',':{w_bas}} "
            f"baslik_en: {ts_metin(e['baslik_en']) + ',':{w_ben}} "
            f"kategori: {ts_metin(e['kategori']) + ',':{w_kat}} "
            f"yil: {e['yil']}, teknik: 'Yağlıboya, tuval üzerine', "
            f"en: {e['en']:>2}, boy: {e['boy']:>2}, "
            f"fiyat: {e['fiyat'] + ',':{w_fiy}} "
            f"durum: {ts_metin(e['durum']) + ',':15} "
            f"ana_fotograf: {ts_metin(e['foto']):{w_foto}}{ek} }},"
        )

    icerik = (
        BASLIK
        + "\n".join(satirlar)
        + ALT
    )
    CIKTI.write_text(icerik, encoding="utf-8")

    satildi = sum(e["durum"] == "satildi" for e in eserler)
    secili = sum(e["secili"] for e in eserler)
    hero = [e["baslik"] for e in eserler if e["hero"]]
    kategoriler = {}
    for e in eserler:
        kategoriler[e["kategori"]] = kategoriler.get(e["kategori"], 0) + 1

    print(f"{CIKTI.relative_to(KOK)} yazıldı.")
    print(f"  Eser        : {len(eserler)}  (yayınlanmayan {atlanan} atlandı)")
    print(f"  Satıldı     : {satildi}   Satışta: {len(eserler) - satildi}")
    print(f"  Seçili      : {secili}")
    print(f"  Hero        : {', '.join(hero) if hero else 'yok'}")
    print(f"  Kategoriler : {', '.join(f'{k} {v}' for k, v in sorted(kategoriler.items()))}")

    # Sessiz veri hatasına karşı basit kontroller
    if satildi == 0 or satildi == len(eserler):
        print("  UYARI: Tüm eserler aynı satış durumunda. 'Durum' sütunu doğru okundu mu?")
    if not hero:
        print("  UYARI: Hiçbir eser hero olarak işaretlenmemiş; ana sayfa ilk eseri kullanır.")
    return 0


BASLIK = """export type Kategori = 'kafeler' | 'sokaklar' | 'figur' | 'portre' | 'diger';
export type Durum    = 'satilabilir' | 'satildi';

export interface Eser {
  slug: string;
  baslik: string;
  baslik_en: string;   // İngilizce sayfalarda kullanılır
  kategori: Kategori;
  yil: number;
  teknik: string;
  en: number;
  boy: number;
  fiyat: number | string; // sayı: TL fiyat | string: galeri notu | 0: satıldı (→ "—")
  durum: Durum;
  ana_fotograf: string;
  aciklama?: string;
  secili?: boolean;     // Ana sayfada seçili eserler bölümünde göster
  hero?: boolean;       // Hero görseli olarak kullan
  yayinlandi?: boolean; // false ise eserler sayfasında gösterme
}

// ---------------------------------------------------------------------------
// BU DOSYA OTOMATİK ÜRETİLİR — elle düzenlemeyin.
// Kaynak: Eserler.xlsx    Üretim: python scripts/eserler-uret.py
// ---------------------------------------------------------------------------
export const eserler: Eser[] = [
"""

ALT = """
];

/** Galeri notu gibi serbest metin fiyatların İngilizce karşılığı. */
const FIYAT_NOTU_EN: Record<string, string> = {
  "Sava Sanat Galerisi'nde Satışta": 'Available at Sava Art Gallery',
};

export function formatFiyat(
  fiyat: number | string,
  durum: Durum,
  dil: 'tr' | 'en' = 'tr',
): string {
  if (durum === 'satildi') return '—';
  if (typeof fiyat === 'string') {
    return dil === 'en' ? (FIYAT_NOTU_EN[fiyat] ?? fiyat) : fiyat;
  }
  if (fiyat <= 0) {
    return dil === 'en' ? 'Price on request' : 'Fiyat için iletişime geçin';
  }
  return '₺' + fiyat.toLocaleString('tr-TR');
}
"""

if __name__ == "__main__":
    raise SystemExit(main())
