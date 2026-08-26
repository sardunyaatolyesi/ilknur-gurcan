# -*- coding: utf-8 -*-
"""
Sergiler.xlsx dosyasından src/data/sergiler.ts üretir.

Kullanım (proje kökünden):
    python scripts/sergiler-uret.py

Excel proje klasörünün bir üstünde beklenir:
    SardunyaWebPage/Sergiler.xlsx

'Göster/Gizle' sütunu 'Gizle' olan sergiler siteye çıkmaz; yalnızca sayılır
ve sayfada "çok sayıda karma sergi" ifadesiyle temsil edilir.
"""
import re
import sys
from datetime import datetime
from pathlib import Path

# Windows konsolu varsayılan olarak cp1254 kullanır; Türkçe olmayan
# karakterlerde (→ gibi) çökmemesi için çıktıyı UTF-8'e sabitle.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl kurulu değil.  Kurulum:  pip install openpyxl")

KOK   = Path(__file__).resolve().parent.parent
XLSX  = KOK.parent / "Sergiler.xlsx"
CIKTI = KOK / "src" / "data" / "sergiler.ts"

AY_TR = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
         "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
AY_EN = ["January", "February", "March", "April", "May", "June",
         "July", "August", "September", "October", "November", "December"]

TUR_EN = {
    "Kişisel": "Solo Exhibition", "Karma": "Group Exhibition",
    "Fuar": "Art Fair", "Festival": "Festival", "Çalıştay": "Workshop",
}

# Sergi ve mekân adlarının İngilizce karşılıkları.
# Listede olmayan bir ad Türkçe hâliyle bırakılır.
AD_EN = {
    "Yılbaşı Sergisi": "New Year Exhibition",
    "İlknur Gürcan Kişisel Sergisi": "İlknur Gürcan Solo Exhibition",
    "12. Art Ankara Çağdaş Sanatlar Fuarı": "12th Art Ankara Contemporary Art Fair",
    "2. Art Ankara Çağdaş Sanatlar Fuarı": "2nd Art Ankara Contemporary Art Fair",
    "2. Uluslararası Bilkent Sanat Festivali": "2nd International Bilkent Art Festival",
    "7. Knidos'un Sır'ı Sanat Çalıştayı": '7th "Secret of Knidos" Art Workshop',
    "RC Sanat Galerisi — Bilkent Sanat Sokağı Açılış":
        "RC Art Gallery — Bilkent Art Street Opening",
    "Kişisel Sergi": "Solo Exhibition",
    "Sardunya Resim Atölyesi Açılış Sergisi": "Sardunya Painting Studio Opening Exhibition",
    "Bilkent 1. Sanat Buluşması": "Bilkent 1st Art Gathering",
}

MEKAN_EN = {
    "Sava Sanat Galerisi": "Sava Art Gallery",
    "ATO - Congresium Kongre ve Sergi Merkezi": "ATO Congresium",
    "RC Sanat Galerisi": "RC Art Gallery",
    "Atilla Sav Sanat Galerisi": "Atilla Sav Art Gallery",
    "Ankara Tenis Kulübü": "Ankara Tennis Club",
    "Sardunya Resim Atölyesi": "Sardunya Painting Studio",
    "Bilkent Sanat Sokağı": "Bilkent Art Street",
    # İngilizcede de aynı kalanlar (uyarı listesinden düşsün diye açıkça yazıldı)
    "Bilkent": "Bilkent",
    "UKKSA": "UKKSA",
}

NOT_EN = {"Mansiyon Ödülü": "Honourable Mention"}


def q(s: str) -> str:
    """TS string literal: kesme işareti varsa çift tırnak."""
    if '"' in s:
        return "'" + s.replace("'", "\\'") + "'"
    return f'"{s}"' if "'" in s else f"'{s}'"


def tarih_coz(deger, satir, alan):
    """'01.12.2026' veya 2018 -> (gun, ay, yil). Gün/ay yoksa None döner."""
    if deger is None:
        return None
    if isinstance(deger, datetime):
        return deger.day, deger.month, deger.year
    # Excel tarihi seri sayı olarak saklamış olabilir (hücre biçimi 'Genel' ise)
    if isinstance(deger, (int, float)) and 20000 < deger < 80000:
        from datetime import timedelta
        d = datetime(1899, 12, 30) + timedelta(days=int(deger))
        print(f"  NOT: satır {satir}, {alan} seri sayı olarak saklanmış "
              f"({int(deger)}) → {d:%d.%m.%Y}. Hücre biçimi tarih yapılabilir.")
        return d.day, d.month, d.year
    s = str(deger).strip()
    if re.fullmatch(r"\d{4}", s):
        return None, None, int(s)
    m = re.fullmatch(r"(\d{1,3})\.(\d{1,2})\.(\d{4})", s)
    if not m:
        print(f"  UYARI: satır {satir}, {alan} okunamadı: {deger!r}")
        return None
    gun, ay, yil = m.groups()
    # Gün alanı üç haneliyse hangi günün kastedildiği belirsizdir
    # ('011' -> 1 mi 11 mi?). Tahmin etmek yerine durdur: yanlış tarih
    # siteye çıkarsa fark edilmesi zor olur.
    if len(gun) > 2:
        sys.exit(f"HATA: Sergiler.xlsx satır {satir}, {alan} = {s!r}\n"
                 f"       Gün alanı üç haneli. Hangi gün kastedildiği belirsiz.\n"
                 f"       Excel'de düzeltip betiği yeniden çalıştırın.")
    if not (1 <= int(gun) <= 31 and 1 <= int(ay) <= 12):
        sys.exit(f"HATA: Sergiler.xlsx satır {satir}, {alan} = {s!r} — geçersiz tarih.")
    return int(gun), int(ay), int(yil)


def tam_tarih(bas, bit, dil):
    """(1,12,2026) ve (31,12,2026) -> '1 Aralık — 31 Aralık 2026'"""
    aylar = AY_TR if dil == "tr" else AY_EN
    if not bas:
        return ""
    if not bas[0]:                          # yalnızca yıl
        return str(bas[2])
    bd, ba, by = bas
    if not bit or not bit[0]:
        return f"{bd} {aylar[ba-1]} {by}"
    ed, ea, ey = bit
    ayrac = " — "
    if by == ey:
        return f"{bd} {aylar[ba-1]}{ayrac}{ed} {aylar[ea-1]} {by}"
    return f"{bd} {aylar[ba-1]} {by}{ayrac}{ed} {aylar[ea-1]} {ey}"


def donem(bas, dil):
    aylar = AY_TR if dil == "tr" else AY_EN
    if not bas:
        return ""
    if not bas[0]:
        return str(bas[2])
    return f"{aylar[bas[1]-1]} {bas[2]}"


def main() -> int:
    if not XLSX.exists():
        sys.exit(f"Excel bulunamadı: {XLSX}")

    ws = openpyxl.load_workbook(XLSX).active
    sut = {}
    for c in range(1, ws.max_column + 1):
        ad = ws.cell(row=1, column=c).value
        if ad and str(ad).strip() not in sut:
            sut[str(ad).strip()] = c

    oku = lambda r, ad: (ws.cell(row=r, column=sut[ad]).value if ad in sut else None)

    yaklasan, gecmis = [], []
    gizli = {}

    for r in range(2, ws.max_row + 1):
        baslik = oku(r, "Sergi Adı")
        if not baslik:
            continue
        baslik = str(baslik).strip()

        if str(oku(r, "Göster/Gizle") or "").strip().lower().startswith("gizle"):
            tur = str(oku(r, "Tür") or "").strip()
            gizli[tur] = gizli.get(tur, 0) + 1
            continue

        bas = tarih_coz(oku(r, "Başlangıç Tarihi"), r, "Başlangıç Tarihi")
        bit = tarih_coz(oku(r, "Bitiş Tarihi"), r, "Bitiş Tarihi")
        tur = str(oku(r, "Tür") or "").strip()
        mekan = str(oku(r, "Mekan") or "").strip()
        sehir = str(oku(r, "Şehir") or "").strip()
        notu = str(oku(r, "Notlar") or "").strip()

        # Görüntülenecek yıl: "2011–12" gibi aralıklar için Dönem sütunu
        kisaDonem = str(oku(r, "Dönem (kısa)") or "").strip()
        yil = kisaDonem if re.fullmatch(r"\d{4}(–\d{2,4})?", kisaDonem) else str(bas[2] if bas else "")

        kayit = {
            "baslik": baslik,
            "baslik_en": AD_EN.get(baslik, baslik),
            "tur": tur,
            "tur_en": TUR_EN.get(tur, tur),
            "mekan": mekan,
            "mekan_en": MEKAN_EN.get(mekan, mekan),
            "sehir": sehir,
            "tarih": tam_tarih(bas, bit, "tr"),
            "tarih_en": tam_tarih(bas, bit, "en"),
            "donem": donem(bas, "tr"),
            "donem_en": donem(bas, "en"),
            "yil": yil,
            "not": notu,
            "not_en": NOT_EN.get(notu, notu),
        }

        if str(oku(r, "Durum") or "").strip().lower().startswith("yaklaş"):
            yaklasan.append(kayit)
        else:
            gecmis.append(kayit)

    yaklasan.sort(key=lambda s: s["yil"])
    gecmis.sort(key=lambda s: s["yil"], reverse=True)

    def blok(ad, liste):
        satirlar = []
        for s in liste:
            alanlar = ", ".join(
                f"{k}: {q(v)}" for k, v in s.items() if v
            )
            satirlar.append(f"  {{ {alanlar} }},")
        return f"export const {ad}: Sergi[] = [\n" + "\n".join(satirlar) + "\n];\n"

    toplamGizli = sum(gizli.values())
    icerik = f"""export interface Sergi {{
  baslik: string;
  baslik_en: string;
  tur: string;
  tur_en: string;
  mekan?: string;
  mekan_en?: string;
  sehir?: string;
  tarih: string;      // "1 Aralık — 31 Aralık 2026"
  tarih_en: string;
  donem: string;      // "Aralık 2026"
  donem_en: string;
  yil: string;        // "2018" veya "2011–12"
  not?: string;
  not_en?: string;
}}

// ---------------------------------------------------------------------------
// BU DOSYA OTOMATİK ÜRETİLİR — elle düzenlemeyin.
// Kaynak: Sergiler.xlsx    Üretim: python scripts/sergiler-uret.py
// ---------------------------------------------------------------------------

{blok("yaklasanSergiler", yaklasan)}
{blok("gecmisSergiler", gecmis)}
/** Listeye alınmayan ('Gizle' işaretli) karma sergi sayısı. */
export const gizliSergiSayisi = {toplamGizli};
"""
    CIKTI.write_text(icerik, encoding="utf-8")

    print(f"{CIKTI.relative_to(KOK)} yazıldı.")
    print(f"  Yaklaşan : {len(yaklasan)}")
    print(f"  Geçmiş   : {len(gecmis)}")
    print(f"  Gizlenen : {toplamGizli}  {gizli if gizli else ''}")
    turler = {}
    for s in gecmis:
        turler[s["tur"]] = turler.get(s["tur"], 0) + 1
    print(f"  Geçmiş tür dağılımı: {turler}")

    # İngilizce karşılığı bulunamayanlar Türkçe kalır; sessizce geçmesin
    cevrilmemis = sorted({s["baslik"] for s in yaklasan + gecmis
                          if s["baslik"] not in AD_EN})
    cevrilmemisMekan = sorted({s["mekan"] for s in yaklasan + gecmis
                               if s["mekan"] and s["mekan"] not in MEKAN_EN})
    if cevrilmemis:
        print(f"  UYARI: İngilizce adı olmayan sergi ({len(cevrilmemis)}): {cevrilmemis}")
        print("         Betikteki AD_EN sözlüğüne eklenmeli; şimdilik Türkçe görünecek.")
    if cevrilmemisMekan:
        print(f"  UYARI: İngilizce adı olmayan mekân ({len(cevrilmemisMekan)}): {cevrilmemisMekan}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
